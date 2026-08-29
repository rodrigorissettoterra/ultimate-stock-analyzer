from __future__ import annotations

import base64
import json
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import httpx
from openpyxl import load_workbook

from ultimate_stock_analyzer.domain.master import SectorClassificationRecord

B3_CLASSIFICATION_APP_URL = (
    "https://sistemaswebb3-listados.b3.com.br/"
    "listedCompaniesPage/classification?language=pt-br"
)
B3_COMPANY_API_BASE = (
    "https://sistemaswebb3-listados.b3.com.br/"
    "listedCompaniesProxy/CompanyCall"
)


@dataclass(frozen=True, slots=True)
class B3IndustryClassificationRow:
    sector: str
    subsector: str
    segment: str
    trading_name: str
    issuer_code: str
    listing_segment: str | None = None


@dataclass(slots=True)
class B3IndustryClassificationCollector:
    user_agent: str = "ultimate-stock-analyzer/0.2"
    timeout_seconds: float = 60.0
    page_size: int = 100
    last_unmapped_issuer_codes: tuple[str, ...] = field(default=(), init=False)

    def _headers(self) -> dict[str, str]:
        return {
            "Accept": "application/json, text/plain, */*",
            "User-Agent": self.user_agent,
            "Referer": B3_CLASSIFICATION_APP_URL,
        }

    def _url(self, endpoint: str, payload: dict[str, object]) -> str:
        encoded = json.dumps(
            payload,
            separators=(",", ":"),
            ensure_ascii=True,
        ).encode()
        token = base64.b64encode(encoded).decode()
        return f"{B3_COMPANY_API_BASE}/{endpoint}/{token}"

    def download_workbook(self) -> bytes:
        payload = {"language": "pt-br"}
        url = self._url("GetDownloadIndustryClassification", payload)
        with httpx.Client(
            timeout=self.timeout_seconds,
            follow_redirects=True,
            headers=self._headers(),
        ) as client:
            session_response = client.get(self._url("GetIndustryClassification", payload))
            session_response.raise_for_status()
            response = client.get(url)
            response.raise_for_status()
        if not response.content.startswith(b"PK"):
            content_type = response.headers.get("content-type", "unknown")
            raise ValueError(
                "B3 industry-classification download is not an XLSX archive: "
                f"content_type={content_type} bytes={len(response.content)}"
            )
        return response.content

    def download_company_catalog_archive(self) -> bytes:
        if not 1 <= self.page_size <= 500:
            raise ValueError("B3 company catalog page_size must be between 1 and 500")

        output = BytesIO()
        expected_total_pages: int | None = None
        expected_total_records: int | None = None
        page_number = 1
        with httpx.Client(
            timeout=self.timeout_seconds,
            follow_redirects=True,
            headers=self._headers(),
        ) as client, ZipFile(output, "w", compression=ZIP_DEFLATED) as archive:
            while True:
                payload = {
                    "language": "pt-br",
                    "pageNumber": page_number,
                    "pageSize": self.page_size,
                }
                response = client.get(self._url("GetInitialCompanies", payload))
                response.raise_for_status()
                body = response.json()
                page = body.get("page") or {}
                results = body.get("results")
                if not isinstance(results, list):
                    raise TypeError("B3 company catalog response has no results list")

                total_pages = _positive_int(page.get("totalPages"), "totalPages")
                total_records = _nonnegative_int(page.get("totalRecords"), "totalRecords")
                returned_page = _positive_int(page.get("pageNumber"), "pageNumber")
                if returned_page != page_number:
                    raise ValueError(
                        "B3 company catalog returned unexpected page number: "
                        f"requested={page_number} returned={returned_page}"
                    )
                if expected_total_pages is None:
                    expected_total_pages = total_pages
                    expected_total_records = total_records
                elif (
                    total_pages != expected_total_pages
                    or total_records != expected_total_records
                ):
                    raise ValueError("B3 company catalog pagination totals changed mid-download")
                if page_number < total_pages and not results:
                    raise ValueError("B3 company catalog returned an empty intermediate page")
                if total_pages > 10_000:
                    raise ValueError("B3 company catalog pagination exceeds safety limit")

                info = ZipInfo(f"page_{page_number:04d}.json")
                info.date_time = (1980, 1, 1, 0, 0, 0)
                info.compress_type = ZIP_DEFLATED
                archive.writestr(info, response.content)
                if page_number >= total_pages:
                    break
                page_number += 1
        return output.getvalue()

    def parse_company_catalog_archive(self, content: bytes) -> list[dict[str, object]]:
        rows: list[dict[str, object]] = []
        expected_total_records: int | None = None
        with ZipFile(BytesIO(content)) as archive:
            names = sorted(
                name
                for name in archive.namelist()
                if name.startswith("page_") and name.endswith(".json")
            )
            if not names:
                raise ValueError("B3 company catalog archive contains no page JSON files")
            for name in names:
                body = json.loads(archive.read(name))
                page = body.get("page") or {}
                results = body.get("results")
                if not isinstance(results, list):
                    raise TypeError(f"B3 company catalog page has no results list: {name}")
                total_records = _nonnegative_int(page.get("totalRecords"), "totalRecords")
                if expected_total_records is None:
                    expected_total_records = total_records
                elif total_records != expected_total_records:
                    raise ValueError("B3 company catalog archive has inconsistent totalRecords")
                rows.extend(dict(item) for item in results if isinstance(item, dict))
        if expected_total_records is None or len(rows) != expected_total_records:
            raise ValueError(
                "B3 company catalog row count mismatch: "
                f"expected={expected_total_records} actual={len(rows)}"
            )
        return rows

    def parse_workbook(self, content: bytes) -> list[B3IndustryClassificationRow]:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        try:
            worksheet = workbook[workbook.sheetnames[0]]
            current_sector: str | None = None
            current_subsector: str | None = None
            current_segment: str | None = None
            by_code: dict[str, B3IndustryClassificationRow] = {}
            for raw_row in worksheet.iter_rows(
                min_col=2,
                max_col=7,
                values_only=True,
            ):
                values = [_text(value) for value in raw_row]
                sector, subsector, segment, trading_name, issuer_code, listing_segment = values
                if _is_header_row(values):
                    continue
                if sector:
                    current_sector = sector
                if subsector:
                    current_subsector = subsector
                if segment:
                    current_segment = segment
                if not issuer_code or not trading_name:
                    continue
                code = issuer_code.upper()
                if not current_sector or not current_subsector or not current_segment:
                    raise ValueError(
                        "B3 industry workbook issuer row has incomplete classification: "
                        f"issuer_code={code}"
                    )
                row = B3IndustryClassificationRow(
                    sector=current_sector,
                    subsector=current_subsector,
                    segment=current_segment,
                    trading_name=trading_name,
                    issuer_code=code,
                    listing_segment=listing_segment,
                )
                existing = by_code.get(code)
                if existing is not None and existing != row:
                    raise ValueError(
                        "B3 industry workbook contains conflicting issuer classification: "
                        f"issuer_code={code}"
                    )
                by_code[code] = row
            if not by_code:
                raise ValueError("B3 industry workbook contains no issuer classifications")
            return list(by_code.values())
        finally:
            workbook.close()

    def normalize(
        self,
        workbook_content: bytes,
        company_catalog_archive: bytes,
        *,
        collected_at: datetime,
    ) -> list[SectorClassificationRecord]:
        workbook_rows = self.parse_workbook(workbook_content)
        catalog_rows = self.parse_company_catalog_archive(company_catalog_archive)
        identity = _active_identity_by_issuer_code(catalog_rows)

        normalized: list[SectorClassificationRecord] = []
        unmapped: list[str] = []
        by_company: dict[str, SectorClassificationRecord] = {}
        for row in workbook_rows:
            issuer_identity = identity.get(row.issuer_code)
            if issuer_identity is None:
                unmapped.append(row.issuer_code)
                continue
            cvm_code, cnpj = issuer_identity
            record = SectorClassificationRecord(
                company_id=f"cvm:{cvm_code}",
                cvm_code=cvm_code,
                cnpj=cnpj,
                issuer_code=row.issuer_code,
                trading_name=row.trading_name,
                sector=row.sector,
                subsector=row.subsector,
                segment=row.segment,
                listing_segment=row.listing_segment,
                collected_at=collected_at,
            )
            existing = by_company.get(record.company_id)
            if existing is not None and (
                existing.sector,
                existing.subsector,
                existing.segment,
            ) != (record.sector, record.subsector, record.segment):
                raise ValueError(
                    "B3 company maps to conflicting economic classifications: "
                    f"company_id={record.company_id}"
                )
            if existing is None:
                by_company[record.company_id] = record
                normalized.append(record)

        self.last_unmapped_issuer_codes = tuple(sorted(set(unmapped)))
        if self.last_unmapped_issuer_codes:
            examples = ", ".join(self.last_unmapped_issuer_codes[:5])
            warnings.warn(
                "Excluding B3 industry-classification rows without an active official "
                "company-catalog identity: "
                f"count={len(self.last_unmapped_issuer_codes)} examples={examples}",
                RuntimeWarning,
                stacklevel=2,
            )
        if not normalized:
            raise ValueError("B3 industry classification produced no normalized companies")
        return normalized


def _active_identity_by_issuer_code(
    rows: list[dict[str, object]],
) -> dict[str, tuple[int, str | None]]:
    grouped: dict[str, list[dict[str, object]]] = {}
    for row in rows:
        if str(row.get("status") or "").strip().upper() != "A":
            continue
        issuer_code = _text(row.get("issuingCompany"))
        cvm_raw = _text(row.get("codeCVM"))
        if not issuer_code or not cvm_raw:
            continue
        grouped.setdefault(issuer_code.upper(), []).append(row)

    output: dict[str, tuple[int, str | None]] = {}
    for issuer_code, items in grouped.items():
        cvm_codes = {int(str(item["codeCVM"]).strip()) for item in items if item.get("codeCVM")}
        if len(cvm_codes) != 1:
            raise ValueError(
                "B3 issuer code maps to multiple active CVM identities: "
                f"issuer_code={issuer_code} cvm_codes={sorted(cvm_codes)}"
            )
        cnpjs = {
            _digits(item.get("cnpj"))
            for item in items
            if _digits(item.get("cnpj")) is not None
        }
        if len(cnpjs) > 1:
            raise ValueError(
                "B3 issuer code maps to multiple active CNPJs: "
                f"issuer_code={issuer_code}"
            )
        output[issuer_code] = (next(iter(cvm_codes)), next(iter(cnpjs), None))
    return output


def _is_header_row(values: list[str | None]) -> bool:
    normalized = {(_plain(value) if value else "") for value in values}
    header_tokens = {
        "setor",
        "setor economico",
        "subsetor",
        "segmento",
        "emissor",
        "nome de pregao",
        "codigo",
        "segmento de negociacao",
    }
    return bool(normalized & header_tokens) and any(
        token in normalized for token in ("codigo", "emissor", "nome de pregao")
    )


def _plain(value: str) -> str:
    import unicodedata

    normalized = unicodedata.normalize("NFKD", value)
    return "".join(ch for ch in normalized if not unicodedata.combining(ch)).casefold().strip()


def _text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _digits(value: object) -> str | None:
    text = _text(value)
    if text is None:
        return None
    digits = "".join(character for character in text if character.isdigit())
    return digits or None


def _positive_int(value: object, field_name: str) -> int:
    parsed = _nonnegative_int(value, field_name)
    if parsed < 1:
        raise ValueError(f"B3 company catalog {field_name} must be positive")
    return parsed


def _nonnegative_int(value: object, field_name: str) -> int:
    if value is None:
        raise ValueError(f"B3 company catalog {field_name} is missing")
    try:
        parsed = int(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"B3 company catalog {field_name} is not an integer") from exc
    if parsed < 0:
        raise ValueError(f"B3 company catalog {field_name} must be non-negative")
    return parsed
